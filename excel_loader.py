import openpyxl
from datetime import datetime, date
from models import db, Client


# Mapeo de columnas del Excel (posición -> campo del modelo)
COLUMN_MAP = {
    1: 'nit_raw',        # NIT (contiene NIT + nombre concatenados)
    2: 'dig_ver',        # DIG.VER.
    3: 'cobrador',       # COBRADOR
    5: 'centro_costo',   # CENTRO DE COSTO
    6: 'subcentro_costo',# SUBCENTRO DE COSTO
    7: 'tercero',        # TERCERO
    8: 'sucursal',       # SUCURS
    9: 'nombre',         # NOMBRE
    10: 'cuenta',        # CUENTA
    11: 'cta',           # CTA
    12: 'descripcion',   # DESCRIPCION
    13: 'saldo',         # SALDO
    14: 'valor_vencido', # VALOR VENCIDO
    15: 'telefono_1',    # TELEFONO-1
    16: 'telefono_2',    # TELEFONO-2
    17: 'ciudad',        # CIUDAD
    18: 'comprobante',   # COMPROBANTE
    19: 'fecha_factura', # FECHA
    20: 'fecha_vencimiento',  # FEC-VENCE
    21: 'rango_360',     # <= 360-
    22: 'rango_359_91',  # DE 359- A 91-
    23: 'rango_90_60',   # DE 90- A 60-
    24: 'rango_59_30',   # DE 59- A 30-
    25: 'rango_29',      # 29- ==>
    26: 'total',         # TOTAL
    27: 'estado',        # ESTADO
    28: 'email',         # EMAIL
}


def parse_date(value):
    """Convierte un valor a fecha."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse_float(value):
    """Convierte un valor a float. Retorna None si está vacío."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Ignorar fórmulas de Excel
        if value.startswith('='):
            return None
        cleaned = value.replace(',', '').strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def has_valid_saldo(value):
    """Verifica si el valor de saldo es válido (no vacío/espacios)."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        cleaned = value.replace(',', '').strip()
        if not cleaned:
            return False
        try:
            float(cleaned)
            return True
        except ValueError:
            return False
    return False


def extract_nit(nit_raw):
    """Extrae el NIT numérico del campo raw que puede contener NIT + nombre."""
    if nit_raw is None:
        return ''
    nit_str = str(nit_raw).strip()
    # Extraer solo los dígitos del inicio
    digits = ''
    for char in nit_str:
        if char.isdigit():
            digits += char
        elif digits:
            break
    return digits if digits else nit_str


def load_excel(filepath, filename):
    """Carga un archivo Excel y retorna los clientes parseados."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    clients_added = 0
    clients_updated = 0
    errors = []

    for row_num in range(8, ws.max_row + 1):
        try:
            row_data = {}
            for col, field in COLUMN_MAP.items():
                row_data[field] = ws.cell(row=row_num, column=col).value

            # Saltar filas sin nombre (filas vacías o de totales)
            nombre = row_data.get('nombre')
            if not nombre or str(nombre).strip() == '':
                continue

            # Saltar filas sin comprobante (no son facturas individuales)
            comprobante_raw = row_data.get('comprobante')
            if not comprobante_raw or str(comprobante_raw).strip() == '':
                continue

            nombre = str(nombre).strip()
            nit = extract_nit(row_data.get('nit_raw'))
            tercero = str(row_data.get('tercero', '')).strip() if row_data.get('tercero') else nit
            comprobante = str(row_data.get('comprobante', '')).strip() if row_data.get('comprobante') else ''

            # Buscar si ya existe (por NIT + comprobante)
            existing = Client.query.filter_by(nit=nit, comprobante=comprobante).first()

            if existing:
                client = existing
                clients_updated += 1
            else:
                client = Client()
                clients_added += 1

            client.nit = nit
            client.dig_ver = int(row_data.get('dig_ver', 0) or 0)
            client.cobrador = str(row_data.get('cobrador', '')).strip() if row_data.get('cobrador') else ''
            client.centro_costo = str(row_data.get('centro_costo', '')).strip() if row_data.get('centro_costo') else ''
            client.subcentro_costo = str(row_data.get('subcentro_costo', '')).strip() if row_data.get('subcentro_costo') else ''
            client.tercero = tercero
            client.sucursal = int(row_data.get('sucursal', 0) or 0)
            client.nombre = nombre
            client.cuenta = str(row_data.get('cuenta', '')).strip() if row_data.get('cuenta') else ''
            client.cta = str(row_data.get('cta', '')).strip() if row_data.get('cta') else ''
            client.descripcion = str(row_data.get('descripcion', '')).strip() if row_data.get('descripcion') else ''
            client.saldo = parse_float(row_data.get('saldo')) or 0.0
            client.valor_vencido = parse_float(row_data.get('valor_vencido')) or 0.0
            client.telefono_1 = str(row_data.get('telefono_1', '')).strip() if row_data.get('telefono_1') else ''
            client.telefono_2 = str(row_data.get('telefono_2', '')).strip() if row_data.get('telefono_2') else ''
            client.ciudad = str(row_data.get('ciudad', '')).strip() if row_data.get('ciudad') else ''
            client.comprobante = comprobante
            client.fecha_factura = parse_date(row_data.get('fecha_factura'))
            client.fecha_vencimiento = parse_date(row_data.get('fecha_vencimiento'))
            client.rango_360 = parse_float(row_data.get('rango_360')) or 0.0
            client.rango_359_91 = parse_float(row_data.get('rango_359_91')) or 0.0
            client.rango_90_60 = parse_float(row_data.get('rango_90_60')) or 0.0
            client.rango_59_30 = parse_float(row_data.get('rango_59_30')) or 0.0
            client.rango_29 = parse_float(row_data.get('rango_29')) or 0.0
            client.total = parse_float(row_data.get('total')) or 0.0
            client.estado = str(row_data.get('estado', 'No Vencido')).strip() if row_data.get('estado') else 'No Vencido'
            client.email = str(row_data.get('email', '')).strip() if row_data.get('email') else ''
            client.archivo_origen = filename

            if not existing:
                db.session.add(client)

        except Exception as e:
            errors.append(f"Fila {row_num}: {str(e)}")

    db.session.commit()
    wb.close()

    return {
        'added': clients_added,
        'updated': clients_updated,
        'errors': errors,
        'total': clients_added + clients_updated,
    }
